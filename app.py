from __future__ import annotations

from pathlib import Path
from datetime import datetime
from typing import List, Dict, Any, Optional

from flask import Flask, render_template, request, jsonify
from openpyxl import Workbook, load_workbook

APP_DIR   = Path(__file__).resolve().parent
DATA_DIR  = APP_DIR / "data"
EXCEL_PATH = DATA_DIR / "atendimentos.xlsx"
SHEET_NAME = "Registros"

# Colunas do Excel (ordem importa)
HEADERS = ["Data/Hora", "PROTOCOLO", "CIRCUITO", "CLIENTE", "SERIAL", "TRATATIVA", "STATUS"]

app = Flask(__name__)


# ─── EXCEL HELPERS ─────────────────────────────────────────────────────────────

def ensure_excel_file() -> None:
    """Garante que existe um Excel com cabeçalho correto (incluindo coluna STATUS)."""
    DATA_DIR.mkdir(parents=True, exist_ok=True)

    if EXCEL_PATH.exists():
        wb = load_workbook(EXCEL_PATH)
        if SHEET_NAME not in wb.sheetnames:
            ws = wb.create_sheet(SHEET_NAME)
            ws.append(HEADERS)
            wb.save(EXCEL_PATH)
            return

        ws = wb[SHEET_NAME]

        # Migração: adiciona coluna STATUS se não existir
        if ws.max_row >= 1:
            header_row = [cell.value for cell in ws[1]]
            if "STATUS" not in header_row:
                col_idx = len(header_row) + 1
                ws.cell(row=1, column=col_idx).value = "STATUS"
                wb.save(EXCEL_PATH)
        else:
            ws.append(HEADERS)
            wb.save(EXCEL_PATH)
        return

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws.append(HEADERS)
    wb.save(EXCEL_PATH)


def append_row(protocolo: str, circuito: str, cliente: str, serial: str,
               tratativa: str, status: str = "") -> None:
    """Adiciona uma nova linha no Excel."""
    wb = load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME]
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws.append([timestamp, protocolo, circuito, cliente, serial, tratativa, status])
    wb.save(EXCEL_PATH)


def _parse_dt(s: str) -> Optional[datetime]:
    s = (s or "").strip()
    if not s:
        return None
    try:
        return datetime.strptime(s, "%Y-%m-%d %H:%M:%S")
    except ValueError:
        return None


def _col_index(ws, name: str) -> Optional[int]:
    """Retorna o índice (1-based) da coluna pelo nome do cabeçalho."""
    for cell in ws[1]:
        if cell.value == name:
            return cell.column
    return None


def read_all_records() -> List[Dict[str, Any]]:
    """Lê o Excel e devolve lista de registros como dicionários."""
    ensure_excel_file()
    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws = wb[SHEET_NAME]

    records: List[Dict[str, Any]] = []

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue

        # Suporte a 6 ou 7 colunas (migração)
        dt_raw    = row[0] if len(row) > 0 else None
        protocolo = row[1] if len(row) > 1 else None
        circuito  = row[2] if len(row) > 2 else None
        cliente   = row[3] if len(row) > 3 else None
        serial    = row[4] if len(row) > 4 else None
        tratativa = row[5] if len(row) > 5 else None
        status    = row[6] if len(row) > 6 else ""

        if isinstance(dt_raw, datetime):
            dt = dt_raw
        else:
            dt = _parse_dt(str(dt_raw)) if dt_raw is not None else None

        records.append({
            "row_index": idx,
            "data_hora": dt.strftime("%Y-%m-%d %H:%M:%S") if dt else (str(dt_raw) if dt_raw else ""),
            "dia":       dt.strftime("%Y-%m-%d") if dt else "",
            "protocolo": str(protocolo or "").strip(),
            "circuito":  str(circuito  or "").strip(),
            "cliente":   str(cliente   or "").strip(),
            "serial":    str(serial    or "").strip(),
            "tratativa": str(tratativa or "").strip(),
            "status":    str(status    or "").strip(),
        })

    records.sort(key=lambda r: r["data_hora"], reverse=True)
    return records


def filter_records(records, date_from, date_to, protocolo, cliente):
    df = (date_from or "").strip()
    dt = (date_to   or "").strip()
    p  = (protocolo or "").strip().lower()
    c  = (cliente   or "").strip().lower()

    dfrom = datetime.strptime(df, "%Y-%m-%d").date() if df else None
    dto   = datetime.strptime(dt, "%Y-%m-%d").date() if dt else None

    out = []
    for r in records:
        if r["dia"]:
            rd = datetime.strptime(r["dia"], "%Y-%m-%d").date()
        else:
            rd = None

        if dfrom and (rd is None or rd < dfrom): continue
        if dto   and (rd is None or rd > dto):   continue
        if p and p not in r["protocolo"].lower(): continue
        if c and c not in r["cliente"].lower():   continue
        out.append(r)

    return out


def count_by_day(records):
    """Conta atendimentos, retidos, campo e % por dia."""
    data: Dict[str, Dict] = {}
    for r in records:
        day = r.get("dia") or ""
        if not day:
            continue
        if day not in data:
            data[day] = {"qtd": 0, "retidos": 0, "campo": 0}
        data[day]["qtd"] += 1
        status = r.get("status", "").upper()
        if "RETIDO" in status:
            data[day]["retidos"] += 1
        elif "CAMPO" in status:
            data[day]["campo"] += 1

    result = []
    for day in sorted(data.keys(), reverse=True):
        d = data[day]
        pct = round(d["retidos"] / d["qtd"] * 100) if d["qtd"] else 0
        result.append({"dia": day, "qtd": d["qtd"],
                        "retidos": d["retidos"], "campo": d["campo"], "pct": pct})
    return result


def count_by_month(records):
    """Consolida atendimentos, retidos, campo e % por mês."""
    data: Dict[str, Dict] = {}
    for r in records:
        day = r.get("dia") or ""
        if not day or len(day) < 7:
            continue
        month = day[:7]   # "YYYY-MM"
        if month not in data:
            data[month] = {"qtd": 0, "retidos": 0, "campo": 0}
        data[month]["qtd"] += 1
        status = r.get("status", "").upper()
        if "RETIDO" in status:
            data[month]["retidos"] += 1
        elif "CAMPO" in status:
            data[month]["campo"] += 1

    result = []
    for month in sorted(data.keys(), reverse=True):
        d = data[month]
        pct = round(d["retidos"] / d["qtd"] * 100) if d["qtd"] else 0
        # Formata "YYYY-MM" → "Mmm/AAAA"
        try:
            from datetime import datetime as dt
            label = dt.strptime(month, "%Y-%m").strftime("%b/%Y").capitalize()
        except Exception:
            label = month
        result.append({"mes": month, "label": label, "qtd": d["qtd"],
                        "retidos": d["retidos"], "campo": d["campo"], "pct": pct})
    return result


# ─── ROTAS ─────────────────────────────────────────────────────────────────────

@app.get("/")
def home():
    ensure_excel_file()
    all_records  = read_all_records()
    today        = datetime.now().strftime("%d/%m/%Y")
    today_iso    = datetime.now().strftime("%Y-%m-%d")
    today_records = [r for r in all_records if r.get("dia") == today_iso]

    # Estatísticas do dia
    retidos  = sum(1 for r in today_records if "RETIDO" in r.get("status", "").upper())
    campo    = sum(1 for r in today_records if "CAMPO" in r.get("status", "").upper())
    total_day = len(today_records)
    pct_retido = round(retidos / total_day * 100, 2) if total_day else 0

    return render_template("index.html",
                           excel_path=str(EXCEL_PATH),
                           today_records=today_records,
                           today_date=today,
                           stats={"retidos": retidos, "campo": campo,
                                  "total": total_day, "pct_retido": pct_retido})


@app.post("/save")
def save():
    ensure_excel_file()
    payload   = request.get_json(silent=True) or {}
    protocolo = (payload.get("protocolo") or "").strip()
    circuito  = (payload.get("circuito")  or "").strip()
    cliente   = (payload.get("cliente")   or "").strip()
    serial    = (payload.get("serial")    or "").strip()
    tratativa = (payload.get("tratativa") or "").strip()
    status    = (payload.get("status")    or "").strip()   # novo campo

    if not tratativa:
        return jsonify({"ok": False, "error": "O campo TRATATIVA está vazio."}), 400

    append_row(protocolo, circuito, cliente, serial, tratativa, status)
    return jsonify({"ok": True, "message": "Registro salvo no Excel com sucesso."})


@app.post("/update")
def update_record():
    ensure_excel_file()
    payload   = request.get_json(silent=True) or {}
    row_index = payload.get("row_index")

    if row_index is None:
        return jsonify({"ok": False, "error": "Índice da linha não informado."}), 400

    protocolo = (payload.get("protocolo") or "").strip()
    circuito  = (payload.get("circuito")  or "").strip()
    cliente   = (payload.get("cliente")   or "").strip()
    serial    = (payload.get("serial")    or "").strip()
    tratativa = (payload.get("tratativa") or "").strip()
    status    = (payload.get("status")    or "").strip()

    if not tratativa:
        return jsonify({"ok": False, "error": "O campo TRATATIVA está vazio."}), 400

    try:
        wb = load_workbook(EXCEL_PATH)
        ws = wb[SHEET_NAME]
        ws.cell(row=row_index, column=2).value = protocolo
        ws.cell(row=row_index, column=3).value = circuito
        ws.cell(row=row_index, column=4).value = cliente
        ws.cell(row=row_index, column=5).value = serial
        ws.cell(row=row_index, column=6).value = tratativa
        ws.cell(row=row_index, column=7).value = status
        wb.save(EXCEL_PATH)
        return jsonify({"ok": True, "message": "Registro atualizado com sucesso."})
    except Exception as e:
        return jsonify({"ok": False, "error": f"Erro ao atualizar: {str(e)}"}), 500


@app.post("/update-status")
def update_status():
    """Atualiza apenas o campo STATUS de uma linha (usado pelos timers/retorno)."""
    ensure_excel_file()
    payload   = request.get_json(silent=True) or {}
    row_index = payload.get("row_index")
    status    = (payload.get("status") or "").strip()

    if row_index is None:
        return jsonify({"ok": False, "error": "Índice da linha não informado."}), 400

    try:
        wb = load_workbook(EXCEL_PATH)
        ws = wb[SHEET_NAME]

        # Garante que a coluna STATUS existe
        status_col = _col_index(ws, "STATUS")
        if status_col is None:
            status_col = ws.max_column + 1
            ws.cell(row=1, column=status_col).value = "STATUS"

        ws.cell(row=row_index, column=status_col).value = status
        wb.save(EXCEL_PATH)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.post("/delete")
def delete_record():
    ensure_excel_file()
    payload   = request.get_json(silent=True) or {}
    row_index = payload.get("row_index")

    if row_index is None:
        return jsonify({"ok": False, "error": "Índice da linha não informado."}), 400

    try:
        wb = load_workbook(EXCEL_PATH)
        ws = wb[SHEET_NAME]
        ws.delete_rows(row_index, 1)
        wb.save(EXCEL_PATH)
        return jsonify({"ok": True, "message": "Registro deletado com sucesso."})
    except Exception as e:
        return jsonify({"ok": False, "error": f"Erro ao deletar: {str(e)}"}), 500


@app.get("/today-records")
def get_today_records():
    ensure_excel_file()
    all_records  = read_all_records()
    today        = datetime.now().strftime("%Y-%m-%d")
    today_records = [r for r in all_records if r.get("dia") == today]

    retidos   = sum(1 for r in today_records if "RETIDO" in r.get("status", "").upper())
    campo     = sum(1 for r in today_records if "CAMPO"  in r.get("status", "").upper())
    total_day = len(today_records)
    pct_retido = round(retidos / total_day * 100) if total_day else 0

    return jsonify({
        "ok": True,
        "records": today_records,
        "stats": {"retidos": retidos, "campo": campo,
                  "total": total_day, "pct_retido": pct_retido}
    })


@app.get("/historico")
def historico():
    ensure_excel_file()
    all_records = read_all_records()

    date_from = request.args.get("date_from", "")
    date_to   = request.args.get("date_to",   "")
    protocolo = request.args.get("protocolo", "")
    cliente   = request.args.get("cliente",   "")

    filtered       = filter_records(all_records, date_from, date_to, protocolo, cliente)
    counts         = count_by_day(filtered)
    monthly_counts = count_by_month(filtered)
    total          = len(filtered)

    retidos = sum(1 for r in filtered if "RETIDO" in r.get("status","").upper())
    campo   = sum(1 for r in filtered if "CAMPO"  in r.get("status","").upper())
    pct     = round(retidos / total * 100, 2) if total else 0

    return render_template(
        "historico.html",
        excel_path=str(EXCEL_PATH),
        records=filtered,
        counts=counts,
        monthly_counts=monthly_counts,
        total=total,
        hist_stats={"retidos": retidos, "campo": campo, "pct": pct},
        filters={"date_from": date_from, "date_to": date_to,
                 "protocolo": protocolo, "cliente": cliente}
    )


if __name__ == "__main__":
    ensure_excel_file()
    app.run(host="127.0.0.1", port=5000, debug=True)
